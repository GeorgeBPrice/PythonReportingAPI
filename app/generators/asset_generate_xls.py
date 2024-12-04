import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def generate_asset_report_excel(data, output_dir="temp_reports"):
    """
    Generate an Excel file for the asset report with totals displayed in a separate table to the right.

    Args:
        data (dict): The report data.
        output_dir (str): Directory to save the generated file.

    Returns:
        str: Path to the generated Excel file.

    Raises:
        ValueError: If data is invalid.
        FileNotFoundError: If an error occurs during file generation.
    """
    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Path for the Excel file
    output_file = os.path.join(output_dir, "AssetReport.xlsx")

    try:
        # Convert data to a Pandas DataFrame
        try:
            df = pd.DataFrame(data)
        except Exception as e:
            raise ValueError(f"Error converting data to DataFrame: {e}")

        # Save the DataFrame to an Excel file
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            try:
                # Write the data to the Excel file
                df.to_excel(writer, index=False, sheet_name="Asset Report")

                # Access the workbook and sheet
                workbook = writer.book
                sheet = writer.sheets["Asset Report"]

                # Determine the starting position for the totals table
                data_start_col = len(df.columns) + 3  # Start 2 columns after the populated data
                totals_start_row = 1  # Start on the second row, assuming headers in row 1

                # Report Title
                sheet.cell(row=totals_start_row, column=data_start_col).value = "Asset Report"
                sheet.cell(row=totals_start_row, column=data_start_col).font = Font(bold=True, size=14)

                # Write totals for each numeric column
                numeric_columns = ["Total Assets", "Active Assets", "Retired Assets", "Total Asset Value"]
                for i, column_name in enumerate(numeric_columns, start=totals_start_row + 1):
                    # Write column name
                    sheet.cell(row=i, column=data_start_col).value = column_name

                    # Write formula for totals
                    try:
                        col_letter = sheet.cell(row=1, column=df.columns.get_loc(column_name) + 1).coordinate[0]
                        sheet.cell(row=i, column=data_start_col + 1).value = f"=SUM({col_letter}2:{col_letter}{len(df) + 1})"
                        sheet.cell(row=i, column=data_start_col + 1).font = Font(bold=True)
                        sheet.cell(row=i, column=data_start_col + 1).alignment = Alignment(horizontal="right")
                    except KeyError:
                        raise ValueError(f"Column '{column_name}' not found in the data.")
            except Exception as e:
                raise ValueError(f"Error generating Excel content: {e}")

        # Adjust column widths for readability
        try:
            workbook = load_workbook(output_file)
            sheet = workbook.active
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = adjusted_width
            workbook.save(output_file)
        except Exception as e:
            raise ValueError(f"Error adjusting column widths: {e}")

    except Exception as e:
        # Clean up the output file if an error occurs
        if os.path.exists(output_file):
            os.remove(output_file)
        raise ValueError(f"Error generating Excel report: {e}")

    # Ensure the output file exists
    if not os.path.exists(output_file):
        raise FileNotFoundError(f"Output file not found: {output_file}")

    return output_file
